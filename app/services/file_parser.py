"""
多格式文档解析器：支持 PDF、Word、Excel、TXT、Markdown、扫描件PDF（OCR）

- 常用格式（PDF/Word/Excel/TXT/MD）依赖轻量库，随运行镜像安装
- 扫描件 OCR（paddleocr + pdf2image/poppler）体积大且需系统依赖，改为"用到再导入"，
  运行镜像不装也能正常跑普通文档，避免拖垮 Docker 镜像体积
"""

import logging
from app.core.log import get_logger
logger = get_logger(__name__)

import os


class FileParser:
    def __init__(self):
        self.ocr = None  # OCR懒加载，用到再初始化，省内存

    def parse_pdf(self, file_path):
        """解析普通PDF（可以复制文字的那种）"""
        from pypdf import PdfReader
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
        return text.strip()

    def _render_pdf_pages(self, file_path, dpi=200):
        """
        把 PDF 每一页渲染成 PIL.Image，返回 list[(image, page_no)]。
        优先用 PyMuPDF（fitz），无需外部 poppler；失败再尝试 pdf2image。
        """
        pages = []
        # 1) PyMuPDF（推荐：纯 Python，无外部二进制依赖）
        try:
            import fitz
            from PIL import Image
            doc = fitz.open(file_path)
            for i, page in enumerate(doc, 1):
                pix = page.get_pixmap(dpi=dpi)
                img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                pages.append((img, i))
            return pages
        except Exception as e:
            logger.info("PyMuPDF 渲染失败（%s），尝试 pdf2image...", e)

        # 2) pdf2image + poppler（旧方案，作为兼容回退）
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(file_path, dpi=dpi)
            for i, img in enumerate(images, 1):
                pages.append((img, i))
            return pages
        except Exception as e:
            logger.warning("pdf2image 渲染也失败（%s）", e)
            return []

    def _ocr_images(self, pages):
        """对 [(PIL.Image, page_no), ...] 逐页 OCR，返回合并文字。"""
        if self.ocr is None:
            from paddleocr import PaddleOCR
            logger.info("正在加载OCR模型...")
            self.ocr = PaddleOCR(use_angle_cls=True, lang="ch", show_log=False)

        full_text = ""
        import numpy as np
        for img, page_no in pages:
            logger.info("OCR识别第 %s/%s 页...", page_no, len(pages))
            try:
                result = self.ocr.ocr(np.array(img), cls=True)
                if result:
                    for line in result:
                        if not line:
                            continue
                        for word_info in line:
                            if word_info and len(word_info) > 1:
                                full_text += word_info[1][0] + "\n"
            except Exception as e:
                logger.warning("第 %s 页 OCR 失败：%s", page_no, e)
        return full_text.strip()

    def _fallback_pdf_meta_text(self, file_path, reason="OCR不可用"):
        """当 PDF 无法提取文字时，返回文件元信息兜底，避免完全空文本。"""
        page_count = 0
        # 先用 pypdf（更可能已安装）获取页数
        try:
            from pypdf import PdfReader
            page_count = len(PdfReader(file_path).pages)
        except Exception:
            # 再尝试 fitz
            try:
                import fitz
                page_count = fitz.open(file_path).page_count
            except Exception:
                pass
        base = os.path.basename(file_path)
        return (
            f"文档《{base}》共 {page_count} 页。"
            f"该 PDF 为扫描件/图片型，当前环境 {reason}，未能识别出具体文字内容。"
            f"若需检索其中文字，请确保已安装 PyMuPDF 与 PaddleOCR，或拉取 Ollama 多模态模型 llava。"
        )

    def parse_scanned_pdf(self, file_path):
        """解析扫描件PDF（图片型，用OCR识别文字）"""
        pages = self._render_pdf_pages(file_path)
        if not pages:
            logger.warning("无法把 PDF 渲染成图片，OCR 跳过，使用元信息兜底")
            return self._fallback_pdf_meta_text(file_path, reason="PDF 渲染失败")

        text = self._ocr_images(pages)
        if text and len(text) >= 10:
            return text

        logger.warning("OCR 未识别到有效文字，使用元信息兜底")
        return self._fallback_pdf_meta_text(file_path, reason="OCR 未识别出文字")

    def parse_txt(self, file_path):
        """解析TXT / Markdown 纯文本文件"""
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read().strip()

    def parse_word(self, file_path):
        """解析Word文档（.docx）"""
        from docx import Document
        doc = Document(file_path)
        text = ""
        for para in doc.paragraphs:
            text += para.text + "\n"
        return text.strip()

    def parse_excel(self, file_path):
        """解析Excel表格"""
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"=== 工作表: {sheet_name} ===\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell else "" for cell in row])
                text += row_text + "\n"
        return text.strip()

    def extract_page_images(self, file_path, out_dir, max_pages=None, dpi=110):
        """
        把 PDF 每页渲染成 PNG 存到 out_dir，返回 [(png_path, page_no), ...]。
        用于「以文搜图」：页面图会落盘并由 image_index 生成描述后入库。
        优先用 PyMuPDF（无需 poppler），失败再回退 pdf2image。
        """
        os.makedirs(out_dir, exist_ok=True)
        result = []

        # 1) PyMuPDF 优先
        try:
            import fitz
            doc = fitz.open(file_path)
            for i, page in enumerate(doc, 1):
                if max_pages and i > max_pages:
                    break
                p = os.path.join(out_dir, f"page_{i}.png")
                try:
                    page.get_pixmap(dpi=dpi).save(p)
                    result.append((p, i))
                except Exception as e:
                    logger.info(" 保存第 %s 页图片失败：%s", i, e)
            if result:
                return result
        except Exception as e:
            logger.info("PyMuPDF 渲染页面图失败（%s），尝试 pdf2image...", e)

        # 2) pdf2image 回退
        try:
            from pdf2image import convert_from_path
            images = convert_from_path(file_path, dpi=dpi)
            for i, img in enumerate(images, 1):
                if max_pages and i > max_pages:
                    break
                p = os.path.join(out_dir, f"page_{i}.png")
                try:
                    img.save(p, "PNG")
                    result.append((p, i))
                except Exception as e:
                    logger.info(" 保存第 %s 页图片失败：%s", i, e)
        except Exception as e:
            logger.info(" PDF 页面渲染失败（%s），跳过图片提取", e)
        return result

    @staticmethod
    def is_image_file(file_path):
        """判断是否为可直接入库的图片文件"""
        return os.path.splitext(file_path)[1].lower() in (".png", ".jpg", ".jpeg")

    def auto_parse(self, file_path):
        """
        自动识别文件格式并解析
        这是对外的主接口，调用这个就行
        """
        ext = os.path.splitext(file_path)[1].lower()  # 取文件后缀

        if ext == ".pdf":
            # 先尝试普通解析，如果提取的文字太少，说明是扫描件，用OCR
            try:
                text = self.parse_pdf(file_path)
            except Exception as e:
                logger.warning("普通 PDF 解析失败（%s），尝试扫描件 OCR...", e)
                text = ""
            if len(text) < 100:  # 文字太少，大概率是扫描件
                logger.info("检测到扫描件PDF，启用OCR识别...")
                text = self.parse_scanned_pdf(file_path)
            return text
        elif ext in [".txt", ".md", ".markdown"]:
            return self.parse_txt(file_path)
        elif ext in [".docx", ".doc"]:
            return self.parse_word(file_path)
        elif ext in [".xlsx", ".xls"]:
            return self.parse_excel(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {ext}")

    def parse_url(self, url, timeout=15):
        """
        网页数据源：抓取 URL 正文并清洗为纯文本
        多源数据流水线的一环——支持把在线文档/文章直接纳入知识库
        """
        import requests
        from bs4 import BeautifulSoup

        headers = {"User-Agent": "Mozilla/5.0 (compatible; RAG-Pipeline/1.0)"}
        resp = requests.get(url, headers=headers, timeout=timeout)
        resp.raise_for_status()
        resp.encoding = resp.apparent_encoding or "utf-8"

        soup = BeautifulSoup(resp.text, "lxml")
        # 去掉脚本/样式/导航等噪声
        for tag in soup(["script", "style", "nav", "footer", "header", "noscript", "form"]):
            tag.decompose()

        # 优先取 <article> / <main>，否则取 body
        main = soup.find("article") or soup.find("main") or soup.body or soup
        text = main.get_text(separator="\n")
        # 压缩多余空行
        lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
        return "\n".join(lines)
